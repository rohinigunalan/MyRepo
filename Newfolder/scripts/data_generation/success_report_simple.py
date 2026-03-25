import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook

def create_myself_reading_success_report():
    print(' Creating Myself Data Reading Success Report...')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    data_file = 'dsr/data/Domestic_Myself_form_data_updated.xlsx'
    output_file = f'dsr/screenshots/Domestic_Myself/Myself_Data_Reading_Success_Report_{timestamp}.xlsx'
    
    os.makedirs('dsr/screenshots/Domestic_Myself', exist_ok=True)
    
    if not os.path.exists(data_file):
        print(f' File not found: {data_file}')
        return
    
    df = pd.read_excel(data_file, engine='openpyxl')
    print(f' Read {len(df)} records from Excel file')
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Myself Data Report'
    
    # Add title
    ws['A1'] = f'Myself Data Reading Success Report - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # Add headers starting from row 3
    for i, col in enumerate(df.columns, 1):
        ws.cell(row=3, column=i, value=str(col))
    
    # Add data starting from row 4
    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value))
    
    wb.save(output_file)
    print(f' Success report saved to: {output_file}')
    print(f' Report contains {len(df)} records')
    return output_file

if __name__ == '__main__':
    create_myself_reading_success_report()
